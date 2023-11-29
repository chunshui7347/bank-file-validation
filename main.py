import sys
import openpyxl

def trim_bank_name(bank_name):
    if bank_name is None:
        return 'AMBANK'

    if bank_name == 'MALAYAN BANKING BERHAD':
        return 'MAYBANK'
    
    bank_name = bank_name.replace(' Bank', '').replace(' Berhad', '').replace(' Islamic', '').replace('(M)', '').replace(' BANK', '').replace(' BERHAD', '').strip()
    return bank_name


def read_xlsx(file_path):
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)

        # Check if the sheets exist in the workbook
        if '1 csv downloaded from bank' not in workbook.sheetnames or '2 excel instruction letter' not in workbook.sheetnames:
            print("Error: One or more required sheets are missing in the workbook.")
            return

        # Get the sheets
        sheet1 = workbook['1 csv downloaded from bank']
        sheet2 = workbook['2 excel instruction letter']

        # Check if the sheets have headers
        headers1 = next(sheet1.iter_rows(min_row=1, max_row=1, values_only=True))
        headers2 = next(sheet2.iter_rows(min_row=1, max_row=1, values_only=True))

        # Get the column indices for the required fields in each sheet
        bene_acc_no_index1 = headers1.index('Bene. Acc. No')
        bene_name_index1 = headers1.index('Bene. Name')
        debit_amount_index1 = headers1.index(' Debit Amount ')
        bene_bank_name_index1 = headers1.index('Bene. Bank Name')

        account_no_index2 = headers2.index('Account No.')
        name_of_payee_index2 = headers2.index('Name of Payee')
        amount_index2 = headers2.index('Amount (RM)')
        bank_name_index2 = headers2.index('Bank Name')

        # Iterate through rows and compare data
        for row1, row2 in zip(sheet1.iter_rows(min_row=2, values_only=True), sheet2.iter_rows(min_row=2, values_only=True)):
            # Extract values for comparison from sheet '1 csv downloaded from bank'
            bene_acc_no1 = row1[bene_acc_no_index1]
            bene_name1 = row1[bene_name_index1]
            debit_amount1 = round(row1[debit_amount_index1], 2)
            bene_bank_name1 = trim_bank_name(row1[bene_bank_name_index1])

            # Extract values for comparison from sheet '2 excel instruction letter'
            account_no2 = row2[account_no_index2]
            name_of_payee2 = row2[name_of_payee_index2]
            amount2 = round(row2[amount_index2], 2)
            bank_name2 = trim_bank_name(row2[bank_name_index2])

            # Compare the information based on the specified conditions
            if bene_acc_no1 and bene_name1 and debit_amount1 and bene_bank_name1 and bene_acc_no1 == account_no2 and bene_name1.upper() == str(name_of_payee2).upper() and debit_amount1 == amount2 and bene_bank_name1.upper() == str(bank_name2).upper():
                print(f"Match found:")
                print(f"- Bene. Acc. No: {bene_acc_no1}, Bene. Name: {bene_name1}, Debit Amount: {debit_amount1}, Bene. Bank Name: {bene_bank_name1}")
                print(f"- Account No.: {account_no2}, Name of Payee: {name_of_payee2}, Amount: {amount2}, Bank Name: {bank_name2}")
            else:
                print(f"Match not found:")
                print(f"- Bene. Acc. No: {bene_acc_no1}, Bene. Name: {bene_name1}, Debit Amount: {debit_amount1}, Bene. Bank Name: {bene_bank_name1}")
                print(f"- Account No.: {account_no2}, Name of Payee: {name_of_payee2}, Amount: {amount2}, Bank Name: {bank_name2}")

    except FileNotFoundError as e:
        print(f"Error: {e}")
    except openpyxl.utils.exceptions.InvalidFileException:
        print(f"Error: {file_path} is not a valid xlsx file.")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    # Check if the script is provided with a command line argument (file path)
    if len(sys.argv) < 2:
        print("Please provide the path of the xlsx file as a command line argument.")
    else:
        file_path = sys.argv[1]
        read_xlsx(file_path)
